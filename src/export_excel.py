"""Export the enriched master CSV to a clean .xlsx that Excel won't mangle.

Opening the CSV directly in Excel strips the leading zero from zip codes and
turns long IDs / phone numbers into scientific notation. This writes an xlsx
where:
  - ID/zip/phone columns are real TEXT cells (leading zeros kept, no sci-notation)
  - dollars stay numeric with thousands separators; counts/flags stay integers
  - match_score shows 3 decimals
  - rows are pre-sorted best-first (Zomma Priority, then Fit, then match_score),
    with each firm's contact rows kept together
  - the header row is bold + frozen, the firm-name column is frozen, and an
    autofilter is applied so the sheet is sortable/filterable on open.

Usage:  python -m src.export_excel
"""
from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

import config

# Must stay TEXT so Excel doesn't strip leading zeros / use scientific notation.
TEXT_COLS = ["crd_number", "sec_number", "office_zip"]
# Integer columns that can contain blanks -> nullable Int (no trailing ".0").
INT_COLS = [
    "employee_count", "investment_advisory_employees", "individual_clients",
    "hnw_clients", "total_accounts", "Zomma Priority", "Zomma Fit",
]
# office_phone is stored as the bare 10-digit number and rendered XXX-XXX-XXXX
# by Excel. International / malformed numbers are kept as their original text.
PHONE_FMT = "000-000-0000"

# Cells starting with these are treated as live formulas by Excel (openpyxl
# stores leading-"=" strings as formula cells; "+", "-", "@" fire when the data
# is opened as CSV). Scraped page text (contact names/titles) is third-party
# controlled, so neutralize at write time — classic spreadsheet-formula injection.
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _defang_cell(v):
    """Prefix an apostrophe so Excel keeps the value as inert text."""
    if isinstance(v, str) and v.startswith(_FORMULA_PREFIXES):
        return "'" + v
    return v


def _normalize_phone(v):
    """Return a US phone as a 10-digit int (drops a leading country-code 1), or
    the original string for international / non-10-digit values, or NA if blank."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return pd.NA
    s = str(v).strip()
    if not s or s.lower() == "nan":
        return pd.NA
    d = re.sub(r"\D", "", s)
    if len(d) == 11 and d.startswith("1"):
        d = d[1:]
    return int(d) if len(d) == 10 else s


def _write_sheet(xl, df: pd.DataFrame, sheet_name: str) -> None:
    """Write one dataframe to a sheet: text-safe IDs, formatted dollars, a bold +
    frozen header (row 1 and the firm-name column), autofilter, capped widths."""
    obj_cols = [c for c in df.columns if df[c].dtype == object]
    if obj_cols:
        df = df.copy()
        for c in obj_cols:
            df[c] = df[c].map(_defang_cell)
    df.to_excel(xl, index=False, sheet_name=sheet_name)
    ws = xl.sheets[sheet_name]
    col_of = {cell.value: cell.column for cell in ws[1]}

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    def set_format(colname: str, number_format: str) -> None:
        if colname not in col_of:
            return
        letter = get_column_letter(col_of[colname])
        for cell in ws[letter][1:]:  # skip header
            cell.number_format = number_format

    for c in TEXT_COLS:
        set_format(c, "@")               # explicit text format
    for c in [x for x in df.columns if "aum" in x.lower() or "dollars" in x.lower()]:
        set_format(c, "#,##0")           # 1,117,783,829
    set_format("match_score", "0.000")
    set_format("office_phone", PHONE_FMT)  # 10-digit number -> 617-951-9969

    for col_name, col_idx in col_of.items():
        letter = get_column_letter(col_idx)
        sample = df[col_name].astype(str).head(400)
        width = max(len(str(col_name)), int(sample.str.len().max() or 0)) + 2
        ws.column_dimensions[letter].width = min(max(width, 10), 45)


def export(csv_path: Path | None = None, xlsx_path: Path | None = None) -> Path:
    csv_path = Path(csv_path) if csv_path else config.latest_ria_master()
    xlsx_path = Path(xlsx_path) if xlsx_path else csv_path.with_suffix(".xlsx")

    # Read ID/zip/phone columns as strings up front so leading zeros / formats survive.
    df = pd.read_csv(csv_path, dtype={c: str for c in TEXT_COLS + ["office_phone"]},
                     low_memory=False)

    # Clean the text columns: blank out missing, strip a stray trailing ".0".
    for c in TEXT_COLS:
        if c in df.columns:
            s = df[c].where(df[c].notna(), "").astype(str).str.strip()
            df[c] = s.str.replace(r"\.0$", "", regex=True).replace({"nan": ""})

    # Normalize phone to a bare 10-digit number (Excel renders the dashes).
    if "office_phone" in df.columns:
        df["office_phone"] = df["office_phone"].map(_normalize_phone)

    # Counts / flags / scores -> nullable Int64 so blanks stay blank, not "1.0".
    int_cols = INT_COLS + [c for c in df.columns if c.startswith("svc_")]
    for c in int_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").astype("Int64")

    # Entity key: RIA masters carry crd_number, FDIC banks cert_number, NCUA
    # credit unions cu_number. First present wins (as in scrape_primary_contact).
    id_col = next((c for c in ("crd_number", "cert_number", "cu_number")
                   if c in df.columns), None)

    # --- pre-sort: best targets first, each firm's contact rows kept together,
    #     and the emailable/named contact floated to the top of the firm block.
    he = df["contact_email"].notna() & (df["contact_email"].astype(str).str.strip() != "")
    hn = df["contact_name"].notna()
    by, asc = [], []
    for col, ascending in [("Zomma Priority", False), ("Zomma Fit", False),
                           ("match_score", False), (id_col, True)]:
        if col is not None and col in df.columns:
            by.append(col)
            asc.append(ascending)
    by += ["_he", "_hn"]
    asc += [False, False]
    df = (df.assign(_he=he.astype(int), _hn=hn.astype(int))
            .sort_values(by=by, ascending=asc, kind="mergesort")
            .drop(columns=["_he", "_hn"])
            .reset_index(drop=True))

    # "Company Only" tab: drop the per-contact columns and collapse to one row
    # per firm. drop_duplicates keeps the first (already-sorted) row, so the
    # firm order matches the master tab exactly. Skipped when no entity-id
    # column exists (nothing to collapse on).
    contact_cols = ["contact_name", "contact_title", "contact_email", "email_source"]
    company = None
    if id_col is not None:
        company = (df.drop(columns=[c for c in contact_cols if c in df.columns])
                     .drop_duplicates(subset=id_col, keep="first")
                     .reset_index(drop=True))

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as xl:
        _write_sheet(xl, df, "RIA Master")
        if company is not None:
            _write_sheet(xl, company, "Company Only")

    print(f"Wrote {xlsx_path}")
    print(f"  RIA Master   : {len(df):,} rows x {df.shape[1]} cols  (one row per firm-contact)")
    if company is not None:
        print(f"  Company Only : {len(company):,} rows x {company.shape[1]} cols  (one row per firm)")
    return xlsx_path


if __name__ == "__main__":
    export()
