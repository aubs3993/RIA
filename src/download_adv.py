"""Stage 1: download the latest SEC Registered Investment Advisers monthly snapshot.

Scrapes the SEC index page for the most recent "Registered Investment Advisers"
zip, downloads it (skipping if cached), unzips, and returns the path to the xlsx.

Heads-up: on the 1st of the month the SEC may post a *partial* preliminary file
(~0.8 MB) days before the full monthly export (~5.2 MB). We download whatever is
newest but warn loudly when the zip is below MIN_EXPECTED_ZIP_BYTES; parse_adv is
the authoritative gate that refuses to act on a partial file (see its docstring).
"""
from __future__ import annotations

import re
import time
import zipfile
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup

import config
from src.utils import get_logger

log = get_logger("download_adv", config.LOG_DIR / "pipeline.log")

# SEC zip names: ia<MMDDYY>.zip  (e.g. ia040126.zip = April 1, 2026).
# Exempt-reporting zips end in -exempt.zip and are excluded here.
_IA_ZIP_RE = re.compile(r"ia(\d{2})(\d{2})(\d{2})\.zip$", re.I)

# Full monthly exports run ~5.2 MB; the SEC's partial month-start file is ~0.8 MB.
# Anything under this threshold is almost certainly the partial file — we warn
# here and let parse_adv reject it on column content (the authoritative check).
MIN_EXPECTED_ZIP_BYTES = 2_000_000


@dataclass
class DownloadResult:
    data_path: Path  # .xlsx or .csv depending on what the SEC ships
    zip_path: Path
    snapshot_label: str  # e.g. "2026-04"
    row_count: int | None
    bytes_downloaded: int

    @property
    def xlsx_path(self) -> Path:  # backwards-compat alias
        return self.data_path


def _fetch_index_html() -> str:
    headers = {"User-Agent": config.USER_AGENT, "Accept": "text/html"}
    resp = requests.get(config.SEC_INDEX_URL, headers=headers, timeout=30)
    resp.raise_for_status()
    return resp.text


def _find_latest_ria_zip(html: str) -> tuple[str, str]:
    """Return (absolute_url, filename) for the most recent RIA zip.

    'Most recent' = the link whose embedded MMDDYY parses to the latest date,
    among links that match ia<MMDDYY>.zip and are NOT exempt zips.
    """
    soup = BeautifulSoup(html, "html.parser")
    candidates: list[tuple[tuple[int, int, int], str, str]] = []
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        fname = href.split("/")[-1].lower()
        if fname.endswith("-exempt.zip"):
            continue
        m = _IA_ZIP_RE.search(fname)
        if not m:
            continue
        mm, dd, yy = (int(m.group(1)), int(m.group(2)), int(m.group(3)))
        # SEC two-digit year: 90+ = 19xx, else 20xx. Snapshots only exist post-2010.
        year = 2000 + yy if yy < 90 else 1900 + yy
        absolute = urljoin(config.SEC_INDEX_URL, href)
        candidates.append(((year, mm, dd), absolute, fname))
    if not candidates:
        raise RuntimeError(
            "No 'Registered Investment Advisers' zip links found on the SEC index page. "
            "The page layout may have changed; inspect SEC_INDEX_URL manually."
        )
    candidates.sort(reverse=True)
    _, url, fname = candidates[0]
    return url, fname


def _snapshot_label(filename: str) -> str:
    """Convert SEC filename ia040126.zip → '2026_04' for our local naming."""
    m = _IA_ZIP_RE.search(filename.lower())
    if not m:
        return "unknown"
    mm, _dd, yy = (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    year = 2000 + yy if yy < 90 else 1900 + yy
    return f"{year}_{mm:02d}"


def _download_zip(url: str, dest: Path) -> int:
    headers = {"User-Agent": config.USER_AGENT}
    log.info("Downloading %s → %s", url, dest)
    with requests.get(url, headers=headers, stream=True, timeout=120) as resp:
        resp.raise_for_status()
        total = 0
        tmp = dest.with_suffix(dest.suffix + ".part")
        with tmp.open("wb") as f:
            for chunk in resp.iter_content(chunk_size=64 * 1024):
                if chunk:
                    f.write(chunk)
                    total += len(chunk)
        tmp.replace(dest)
    return total


def _unzip_data_file(zip_path: Path, out_dir: Path) -> Path:
    """Extract the main data file from the SEC zip. Accepts .xlsx or .csv —
    the SEC has shipped both over time."""
    with zipfile.ZipFile(zip_path) as zf:
        members = [n for n in zf.namelist() if n.lower().endswith((".xlsx", ".csv"))]
        if not members:
            raise RuntimeError(f"No .xlsx or .csv inside {zip_path.name}")
        # Pick the largest — the SEC ships one main file per zip.
        members.sort(key=lambda n: zf.getinfo(n).file_size, reverse=True)
        member = members[0]
        target = out_dir / Path(member).name
        if target.exists() and target.stat().st_size == zf.getinfo(member).file_size:
            log.info("Data file already extracted: %s", target)
            return target
        with zf.open(member) as src, target.open("wb") as dst:
            dst.write(src.read())
        log.info("Extracted %s (%.1f MB)", target.name, target.stat().st_size / 1e6)
        return target


def _quick_row_count(data_path: Path) -> int | None:
    """Best-effort row count for either .xlsx or .csv. None on failure."""
    try:
        if data_path.suffix.lower() == ".csv":
            with data_path.open("r", encoding="utf-8", errors="replace", newline="") as f:
                # Subtract 1 for header. Streaming line count avoids loading whole file.
                n = sum(1 for _ in f)
            return max(n - 1, 0)
        from openpyxl import load_workbook

        wb = load_workbook(data_path, read_only=True, data_only=True)
        ws = wb.active
        n = sum(1 for _ in ws.iter_rows(values_only=True))
        wb.close()
        return max(n - 1, 0)
    except Exception as exc:  # pragma: no cover - diagnostic only
        log.warning("Could not row-count %s: %s", data_path.name, exc)
        return None


def download_latest(force: bool = False) -> DownloadResult:
    config.ensure_dirs()

    html = _fetch_index_html()
    url, fname = _find_latest_ria_zip(html)
    label = _snapshot_label(fname)
    zip_dest = config.RAW_DIR / f"ia_{label}.zip"

    if zip_dest.exists() and not force:
        bytes_dl = 0
        log.info("Zip already cached: %s (%.1f MB) — skipping download", zip_dest.name, zip_dest.stat().st_size / 1e6)
    else:
        # Be polite even on the index → download hop.
        time.sleep(1.0)
        bytes_dl = _download_zip(url, zip_dest)

    zip_bytes = zip_dest.stat().st_size
    if zip_bytes < MIN_EXPECTED_ZIP_BYTES:
        log.warning(
            "%s is only %.1f MB (< %.1f MB expected) — this is almost certainly the "
            "SEC's PARTIAL month-start file, not the full export. parse_adv will refuse "
            "to act on it if the Item-5 columns are missing.",
            zip_dest.name, zip_bytes / 1e6, MIN_EXPECTED_ZIP_BYTES / 1e6,
        )

    data_path = _unzip_data_file(zip_dest, config.RAW_DIR)
    rows = _quick_row_count(data_path)

    print(
        f"[download_adv] snapshot={label}  zip={zip_dest.name} "
        f"({zip_dest.stat().st_size / 1e6:.1f} MB)  data={data_path.name}  "
        f"rows≈{rows if rows is not None else '?'}"
    )

    return DownloadResult(
        data_path=data_path,
        zip_path=zip_dest,
        snapshot_label=label,
        row_count=rows,
        bytes_downloaded=bytes_dl,
    )


if __name__ == "__main__":
    download_latest()
